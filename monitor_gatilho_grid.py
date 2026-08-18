#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Monitor de gatilho para robô de grid na Bybit.

Lê uma planilha com setups de interesse, monitora o preço atual na Bybit e envia
alerta via Telegram quando o preço entra na tolerância configurada em relação ao
GATILHO. O script NÃO cria ordens nem configura robô na Bybit; apenas avisa.

Planilha esperada:
- Arquivo padrão: monitor_gatilho_grid.xlsx
- Aba padrão: MONITOR
- Uma linha por setup a monitorar
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, Optional, Tuple

import requests
from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
except Exception:  # pragma: no cover
    load_dotenv = None

try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover
    ZoneInfo = None


BYBIT_TICKERS_URL = "https://api.bybit.com/v5/market/tickers"
TELEGRAM_SEND_MESSAGE_URL = "https://api.telegram.org/bot{token}/sendMessage"

STATUS_ATIVO = "ATIVO"
STATUS_ALERTADO = "ALERTADO"
STATUS_EXPIRADO = "EXPIRADO"
STATUS_CANCELADO = "CANCELADO"

# Cabeçalhos esperados na aba MONITOR.
# Campos usados na lógica: ATIVO, MERCADO, GATILHO, TOLERANCIA_PCT, VALIDADE_ATE, STATUS.
# Campos preenchidos pelo script: STATUS, PRECO_ALERTA, DISTANCIA_ALERTA_PCT, ALERTADO_EM, OBSERVACAO.
# Campos informativos: os demais entram apenas na mensagem do Telegram.
HEADERS = [
    "ATIVO",
    "TIMEFRAME",
    "MERCADO",
    "TIME_STAMP",
    "SETUP",
    "DIRECAO",
    "TOLERANCIA_PCT",
    "VALIDADE_ATE",
    "FAIXA_MIN",
    "FAIXA_MAX",
    "NUM_GRIDS",
    "TRAILING_STOP",
    "GATILHO",
    "TAKE_PROFIT",
    "STOP_LOSS",
    "TP/SL",
    "STATUS",
    "PRECO_ALERTA",
    "DISTANCIA_ALERTA_PCT",
    "ALERTADO_EM",
    "OBSERVACAO",
]

REQUIRED_HEADERS = [
    "ATIVO",
    "MERCADO",
    "GATILHO",
    "TOLERANCIA_PCT",
    "VALIDADE_ATE",
    "STATUS",
]

OUTPUT_HEADERS = ["PRECO_ALERTA", "DISTANCIA_ALERTA_PCT", "ALERTADO_EM", "OBSERVACAO"]


@dataclass
class MonitorConfig:
    arquivo: Path
    aba: str
    intervalo_segundos: int
    timezone: str
    telegram_token: Optional[str]
    telegram_chat_id: Optional[str]
    dry_run: bool = False
    once: bool = False


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )


def normalizar_header(value: Any) -> str:
    return str(value or "").strip().upper()


def normalizar_status(value: Any) -> str:
    return str(value or "").strip().upper()


def get_tz(tz_name: str):
    if ZoneInfo is None:
        return None
    try:
        return ZoneInfo(tz_name)
    except Exception:
        logging.warning("Timezone inválido '%s'. Usando horário local do sistema.", tz_name)
        return None


def now_in_timezone(tz_name: str) -> datetime:
    tz = get_tz(tz_name)
    if tz is None:
        return datetime.now()
    return datetime.now(tz)


def parse_datetime(value: Any, tz_name: str) -> Optional[datetime]:
    """Interpreta datas vindas do Excel ou texto."""
    if value is None or str(value).strip() == "":
        return None

    tz = get_tz(tz_name)

    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        formats = [
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d/%m/%y %H:%M",
        ]
        dt = None
        for fmt in formats:
            try:
                dt = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if dt is None:
            raise ValueError(f"Data/hora inválida: {value!r}")

    if tz is not None and dt.tzinfo is None:
        dt = dt.replace(tzinfo=tz)
    return dt


def to_float(value: Any, field_name: str) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError(f"Campo obrigatório vazio: {field_name}")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "")
    # Aceita tanto 220.30 quanto 220,30.
    if "," in text and "." in text:
        # Ex.: 1.234,56 -> 1234.56
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    return float(text)


def fmt_value(value: Any) -> str:
    if value is None or str(value).strip() == "":
        return "-"
    if isinstance(value, float):
        return f"{value:.8f}".rstrip("0").rstrip(".").replace(".", ",")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    return str(value)


def fmt_pct(value: float) -> str:
    return f"{value:.2f}%".replace(".", ",")


def ensure_headers(ws) -> Dict[str, int]:
    """Garante colunas de saída e devolve mapa header -> coluna."""
    header_map: Dict[str, int] = {}

    for col in range(1, ws.max_column + 1):
        header = normalizar_header(ws.cell(row=1, column=col).value)
        if header:
            header_map[header] = col

    # Cria colunas de saída se faltarem.
    next_col = ws.max_column + 1
    for header in OUTPUT_HEADERS:
        if header not in header_map:
            ws.cell(row=1, column=next_col).value = header
            header_map[header] = next_col
            next_col += 1

    missing = [h for h in REQUIRED_HEADERS if h not in header_map]
    if missing:
        raise RuntimeError(f"Cabeçalhos obrigatórios ausentes na aba {ws.title}: {', '.join(missing)}")

    return header_map


def row_dict(ws, row: int, header_map: Dict[str, int]) -> Dict[str, Any]:
    return {header: ws.cell(row=row, column=col).value for header, col in header_map.items()}


def iter_active_rows(ws, header_map: Dict[str, int]) -> Iterable[Tuple[int, Dict[str, Any]]]:
    for row in range(2, ws.max_row + 1):
        values = row_dict(ws, row, header_map)
        ativo = str(values.get("ATIVO") or "").strip()
        if not ativo:
            continue
        yield row, values


def get_bybit_last_price(symbol: str, category: str = "linear", timeout: int = 10) -> float:
    params = {"category": category, "symbol": symbol}
    response = requests.get(BYBIT_TICKERS_URL, params=params, timeout=timeout)
    response.raise_for_status()
    data = response.json()

    if data.get("retCode") != 0:
        raise RuntimeError(f"Bybit retCode={data.get('retCode')} retMsg={data.get('retMsg')}")

    items = data.get("result", {}).get("list", [])
    if not items:
        raise RuntimeError(f"Ticker não encontrado na Bybit: {category} {symbol}")

    last_price = items[0].get("lastPrice")
    if last_price is None:
        raise RuntimeError(f"Resposta da Bybit sem lastPrice para {symbol}")

    return float(last_price)


def send_telegram_message(token: str, chat_id: str, text: str, timeout: int = 10) -> None:
    url = TELEGRAM_SEND_MESSAGE_URL.format(token=token)
    payload = {"chat_id": chat_id, "text": text}
    response = requests.post(url, data=payload, timeout=timeout)
    response.raise_for_status()
    data = response.json()
    if not data.get("ok"):
        raise RuntimeError(f"Telegram retornou erro: {data}")


def build_alert_message(values: Dict[str, Any], preco_atual: float, distancia_pct: float) -> str:
    validade = values.get("VALIDADE_ATE")
    if isinstance(validade, datetime):
        validade_txt = validade.strftime("%d/%m/%Y %H:%M")
    else:
        validade_txt = fmt_value(validade)

    time_stamp = values.get("TIME_STAMP")
    if isinstance(time_stamp, datetime):
        time_stamp_txt = time_stamp.strftime("%d/%m/%Y %H:%M")
    else:
        time_stamp_txt = fmt_value(time_stamp)

    lines = [
        "🚨 Gatilho dentro do range Bybit",
        "",
        f"Ativo: {fmt_value(values.get('ATIVO'))}",
        f"Timeframe: {fmt_value(values.get('TIMEFRAME'))}",
        f"Mercado: {fmt_value(values.get('MERCADO'))}",
        f"Time stamp: {time_stamp_txt}",
        f"Setup: {fmt_value(values.get('SETUP'))}",
        f"Direção: {fmt_value(values.get('DIRECAO'))}",
        "",
        f"Gatilho: {fmt_value(values.get('GATILHO'))}",
        f"Preço atual: {fmt_value(preco_atual)}",
        f"Distância do gatilho: {fmt_pct(distancia_pct)}",
        f"Validade do setup: {validade_txt}",
        "",
        "Dados para configurar o Grid:",
        f"Faixa: {fmt_value(values.get('FAIXA_MIN'))} até {fmt_value(values.get('FAIXA_MAX'))}",
        f"Número de grids: {fmt_value(values.get('NUM_GRIDS'))}",
        f"Trailing Stop: {fmt_value(values.get('TRAILING_STOP'))}",
        f"Take Profit: {fmt_value(values.get('TAKE_PROFIT'))}",
        f"Stop Loss: {fmt_value(values.get('STOP_LOSS'))}",
        f"TP/SL: {fmt_value(values.get('TP/SL'))}",
        "",
        "Já é possível tentar configurar o robô na Bybit.",
    ]
    return "\n".join(lines)


def process_once(config: MonitorConfig) -> Tuple[int, int, int]:
    """Processa uma rodada. Retorna (ativos, alertados, expirados)."""
    if not config.arquivo.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {config.arquivo}")

    wb = load_workbook(config.arquivo)
    if config.aba not in wb.sheetnames:
        raise RuntimeError(f"Aba '{config.aba}' não encontrada em {config.arquivo}")

    ws = wb[config.aba]
    header_map = ensure_headers(ws)

    agora = now_in_timezone(config.timezone)
    total_ativos = 0
    total_alertados = 0
    total_expirados = 0
    houve_alteracao = False

    for row, values in iter_active_rows(ws, header_map):
        status = normalizar_status(values.get("STATUS"))
        if status != STATUS_ATIVO:
            continue

        total_ativos += 1
        ativo = str(values.get("ATIVO") or "").strip().upper()
        mercado = str(values.get("MERCADO") or "linear").strip().lower()

        try:
            validade = parse_datetime(values.get("VALIDADE_ATE"), config.timezone)
            if validade is not None and agora > validade:
                ws.cell(row=row, column=header_map["STATUS"]).value = STATUS_EXPIRADO
                ws.cell(row=row, column=header_map["OBSERVACAO"]).value = f"Expirado em {agora.strftime('%d/%m/%Y %H:%M:%S')}"
                total_expirados += 1
                houve_alteracao = True
                logging.info("[%s] Expirado pela validade.", ativo)
                continue

            gatilho = to_float(values.get("GATILHO"), "GATILHO")
            tolerancia_pct = to_float(values.get("TOLERANCIA_PCT"), "TOLERANCIA_PCT")
            preco_atual = get_bybit_last_price(ativo, mercado)
            distancia_pct = abs(preco_atual - gatilho) / gatilho * 100.0

            logging.info(
                "[%s] preço=%s gatilho=%s distância=%.4f%% tolerância=%.4f%%",
                ativo,
                preco_atual,
                gatilho,
                distancia_pct,
                tolerancia_pct,
            )

            if distancia_pct <= tolerancia_pct:
                message = build_alert_message(values, preco_atual, distancia_pct)

                if config.dry_run:
                    logging.info("[DRY-RUN] Mensagem Telegram:\n%s", message)
                else:
                    if not config.telegram_token or not config.telegram_chat_id:
                        raise RuntimeError(
                            "TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID não configurados no .env ou linha de comando."
                        )
                    send_telegram_message(config.telegram_token, config.telegram_chat_id, message)

                ws.cell(row=row, column=header_map["STATUS"]).value = STATUS_ALERTADO
                ws.cell(row=row, column=header_map["PRECO_ALERTA"]).value = preco_atual
                ws.cell(row=row, column=header_map["DISTANCIA_ALERTA_PCT"]).value = round(distancia_pct, 4)
                ws.cell(row=row, column=header_map["ALERTADO_EM"]).value = agora.strftime("%d/%m/%Y %H:%M:%S")
                ws.cell(row=row, column=header_map["OBSERVACAO"]).value = "Alerta enviado pelo monitor"
                total_alertados += 1
                houve_alteracao = True
                logging.info("[%s] ALERTADO. Distância %.4f%% <= %.4f%%", ativo, distancia_pct, tolerancia_pct)

        except Exception as exc:
            ws.cell(row=row, column=header_map["OBSERVACAO"]).value = f"Erro: {exc}"
            houve_alteracao = True
            logging.exception("Erro ao processar linha %s (%s): %s", row, ativo, exc)

    if houve_alteracao:
        wb.save(config.arquivo)
        logging.info("Planilha atualizada: %s", config.arquivo)

    return total_ativos, total_alertados, total_expirados


def parse_args() -> MonitorConfig:
    parser = argparse.ArgumentParser(description="Monitor de gatilho para grid Bybit com alerta Telegram")
    parser.add_argument("--arquivo", default="monitor_gatilho_grid.xlsx", help="Planilha de monitoramento")
    parser.add_argument("--aba", default="MONITOR", help="Aba da planilha")
    parser.add_argument("--intervalo-segundos", type=int, default=30, help="Intervalo entre consultas")
    parser.add_argument("--timezone", default="America/Sao_Paulo", help="Timezone para validade e logs")
    parser.add_argument("--once", action="store_true", help="Executa uma rodada e encerra")
    parser.add_argument("--dry-run", action="store_true", help="Não envia Telegram; apenas registra no log")
    parser.add_argument("--telegram-token", default=None, help="Token do bot Telegram; se omitido, usa .env")
    parser.add_argument("--telegram-chat-id", default=None, help="Chat ID Telegram; se omitido, usa .env")

    args = parser.parse_args()

    if load_dotenv is not None:
        load_dotenv()

    token = args.telegram_token or os.environ.get("TELEGRAM_BOT_TOKEN") or os.environ.get("BOT_TOKEN")
    chat_id = args.telegram_chat_id or os.environ.get("TELEGRAM_CHAT_ID") or os.environ.get("CHAT_ID")

    return MonitorConfig(
        arquivo=Path(args.arquivo),
        aba=args.aba,
        intervalo_segundos=max(5, args.intervalo_segundos),
        timezone=args.timezone,
        telegram_token=token,
        telegram_chat_id=chat_id,
        dry_run=args.dry_run,
        once=args.once,
    )


def main() -> int:
    setup_logging()
    config = parse_args()

    logging.info("Monitor iniciado | arquivo=%s | aba=%s | intervalo=%ss", config.arquivo, config.aba, config.intervalo_segundos)

    while True:
        try:
            ativos, alertados, expirados = process_once(config)
            logging.info("Rodada concluída | ativos=%s | alertados=%s | expirados=%s", ativos, alertados, expirados)
        except KeyboardInterrupt:
            logging.info("Monitor interrompido pelo usuário.")
            return 0
        except Exception as exc:
            logging.exception("Erro na rodada do monitor: %s", exc)

        if config.once:
            break

        time.sleep(config.intervalo_segundos)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
